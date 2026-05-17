"""Viega Viptool Master Excel adapter.

Phase 11.2: pragmatischer "informierter" Stub mit Default-Spaltenmapping.

Solange keine echte Viptool-Beispieldatei vorliegt, basiert das Mapping auf
oeffentlicher Doku zum Viega-Viptool-Master-Export. Sobald ein realer Export
verfuegbar ist, werden die Spaltennamen / Einheiten verfeinert. Bis dahin
ist dieser Adapter NICHT der Default fuer beliebige xlsx-Dateien — er
claimt nur Dateien, deren Name "viptool" enthaelt oder in deren ersten
Zellen ein Hinweis auf Viega/Viptool steht. Alle Dateien, die nicht
eindeutig zugeordnet werden koennen, gehen weiterhin an den GenericTableImporter.
"""
from __future__ import annotations

import io
import re
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.models.heating import (
    HeatingCircuitBase,
    HeatingDesignBase,
    HeatingDesignImportPreview,
)
from app.services.heating_importers.base import (
    ColumnMapping,
    HeatingImporter,
    HeatingImporterError,
)


# ---------------------------------------------------------------------------
# Default column mapping (canonical field -> list of possible header tokens)
#
# Quelle: oeffentlich verfuegbare Doku/Screenshots zum Viega Viptool Master
# Export. Tokens werden case-insensitive und ohne diakritische Zeichen mit
# Header-Zellen verglichen; Treffer per Substring (fuzzy genug fuer
# typische Spaltenvarianten wie "Norm-Heizlast Q [W]" oder "Volumenstrom kg/h").
# ---------------------------------------------------------------------------
VIPTOOL_DEFAULT_MAPPING: dict[str, list[str]] = {
    "strand": ["strang", "bauteil"],
    "room": ["raum", "bezeichnung"],
    "floor": ["geschoss", "ebene"],
    "radiator_type": ["heizflaeche", "heizflache", "heizkoerper", "heizkorper", "heizflache"],
    "heat_load_w": ["norm-heizlast", "heizlast", "qn", "q"],
    "volume_flow_lph": ["volumenstrom", "massenstrom", "durchfluss"],
    "pressure_drop_pa": ["druckverlust", "delta p", "dp"],
    "pipe_length_m": ["rohrlaenge", "rohrlange", "laenge", "lange"],
    "valve_type": ["ventil", "ventiltyp"],
    "valve_preset": ["voreinstellung", "preset", "stellwert"],
    "kv_value": ["kv", "kv-wert", "kvs"],
}


# Welche Felder sind reine Substring-Treffer und welche brauchen einen
# exakten Token-Vergleich? "q" und "kv" sind kurz und triggern sonst zu
# leicht False Positives ("Quelle", "Skv", ...).
_EXACT_TOKEN_FIELDS = {"heat_load_w": {"qn", "q"}, "kv_value": {"kv", "kvs"}}


_WORKSHEET_HINTS = ("strang", "heizk", "kreis", "berechnung")


def _normalise(text: str) -> str:
    """Lowercase, Umlaute weg, Whitespace zu einzelnem Space."""
    lowered = text.lower().strip()
    replacements = {
        "ae": "ae",  # placeholder; explicit umlaute next
        "ä": "ae",
        "ö": "oe",
        "ü": "ue",
        "ß": "ss",
    }
    for src, dst in replacements.items():
        lowered = lowered.replace(src, dst)
    lowered = re.sub(r"\s+", " ", lowered)
    return lowered


def _header_matches(header_cell: str, tokens: Iterable[str], *, exact_tokens: set[str] | None = None) -> bool:
    norm = _normalise(header_cell)
    exact_tokens = exact_tokens or set()
    for token in tokens:
        tok = _normalise(token)
        if not tok:
            continue
        if tok in exact_tokens:
            # exact match against tokenised header words
            words = re.findall(r"[a-z0-9]+", norm)
            if tok in words:
                return True
        else:
            if tok in norm:
                return True
    return False


# ---------------------------------------------------------------------------
# Unit detection + conversion helpers
# ---------------------------------------------------------------------------


_UNIT_PATTERN = re.compile(r"[\[\(]\s*([^\]\)]+?)\s*[\]\)]")


def _detect_unit(header_cell: str) -> str | None:
    match = _UNIT_PATTERN.search(header_cell)
    if match:
        return match.group(1).strip().lower()
    return None


def _to_watts(value: float, unit: str | None) -> float:
    if unit is None:
        return value
    u = unit.lower()
    if "kw" in u:
        return value * 1000.0
    return value  # default: already W


def _to_pascal(value: float, unit: str | None) -> float:
    if unit is None:
        return value
    u = unit.lower()
    if "mbar" in u:
        return value * 100.0
    if u.strip() == "bar":
        return value * 100000.0
    if "kpa" in u:
        return value * 1000.0
    return value  # default: already Pa


def _to_lph(value: float, unit: str | None) -> float:
    # kg/h ~ l/h for water at typical heating temperatures (rho ~ 0.97-1.0).
    # m^3/h -> l/h *1000.
    if unit is None:
        return value
    u = unit.lower()
    if "m3/h" in u or "m³/h" in u:
        return value * 1000.0
    return value


def _to_meters(value: float, unit: str | None) -> float:
    if unit is None:
        return value
    u = unit.lower()
    if u.strip() == "cm":
        return value / 100.0
    if u.strip() == "mm":
        return value / 1000.0
    return value


def _coerce_float(raw: object) -> float | None:
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    if isinstance(raw, str):
        s = raw.strip().replace(",", ".")
        if not s:
            return None
        # extract first float-looking token (handles "1234 W")
        m = re.search(r"-?\d+(?:\.\d+)?", s)
        if not m:
            return None
        try:
            return float(m.group(0))
        except ValueError:
            return None
    return None


def _coerce_str(raw: object) -> str | None:
    if raw is None:
        return None
    s = str(raw).strip()
    return s or None


# ---------------------------------------------------------------------------
# Worksheet / header detection
# ---------------------------------------------------------------------------


def _pick_worksheet(wb) -> Worksheet:
    """Prefer worksheets with viptool-typical names; fall back to the first non-empty one."""
    for ws in wb.worksheets:
        name = _normalise(ws.title or "")
        if any(hint in name for hint in _WORKSHEET_HINTS):
            return ws
    # Fallback: first sheet with any content.
    for ws in wb.worksheets:
        if ws.max_row and ws.max_column:
            return ws
    raise HeatingImporterError("Excel-Datei enthaelt keine lesbaren Tabellenblaetter.")


def _find_header_row(ws: Worksheet, *, max_scan_rows: int = 25) -> tuple[int, list[str]]:
    """Return (row_index_1_based, header_values) — the first row that matches
    at least 3 distinct canonical fields from the default mapping."""
    best_row: int | None = None
    best_headers: list[str] = []
    best_hits = 0

    row_limit = min(max_scan_rows, ws.max_row or 0)
    for row_idx in range(1, row_limit + 1):
        row_values = [
            _coerce_str(cell.value) or ""
            for cell in ws[row_idx]
        ]
        if not any(row_values):
            continue
        hits = 0
        for canonical, tokens in VIPTOOL_DEFAULT_MAPPING.items():
            exact = _EXACT_TOKEN_FIELDS.get(canonical)
            for cell_text in row_values:
                if cell_text and _header_matches(cell_text, tokens, exact_tokens=exact):
                    hits += 1
                    break
        if hits > best_hits:
            best_hits = hits
            best_row = row_idx
            best_headers = row_values

    if best_row is None or best_hits < 3:
        raise HeatingImporterError(
            "Keine Viptool-konforme Kopfzeile gefunden. "
            "Mindestens 3 erkennbare Spalten (z.B. Strang, Heizlast, Volumenstrom) erforderlich."
        )
    return best_row, best_headers


def _build_column_mapping(headers: list[str]) -> tuple[dict[str, int], dict[str, str], dict[str, str | None]]:
    """Map canonical field -> column index, plus column-label-by-canonical
    and unit-by-canonical (parsed from header brackets).
    """
    by_field: dict[str, int] = {}
    labels: dict[str, str] = {}
    units: dict[str, str | None] = {}
    for canonical, tokens in VIPTOOL_DEFAULT_MAPPING.items():
        exact = _EXACT_TOKEN_FIELDS.get(canonical)
        for col_idx, header in enumerate(headers):
            if not header:
                continue
            if _header_matches(header, tokens, exact_tokens=exact):
                if canonical not in by_field:
                    by_field[canonical] = col_idx
                    labels[canonical] = header
                    units[canonical] = _detect_unit(header)
                break
    return by_field, labels, units


# ---------------------------------------------------------------------------
# Brand detection in first row
# ---------------------------------------------------------------------------


def _first_cells_mention_brand(content: bytes) -> bool:
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        return False
    try:
        for ws in wb.worksheets:
            row_iter = ws.iter_rows(min_row=1, max_row=1, max_col=5, values_only=True)
            for row in row_iter:
                for cell in row:
                    if cell is None:
                        continue
                    txt = _normalise(str(cell))
                    if "viega" in txt or "viptool" in txt:
                        return True
            break  # only first sheet
    finally:
        wb.close()
    return False


# ---------------------------------------------------------------------------
# Importer implementation
# ---------------------------------------------------------------------------


class ViptoolMasterImporter(HeatingImporter):
    """Viega Viptool Master Excel-Export -> internes heating_design-Schema."""

    source_name = "viptool_xlsx"
    display_name = "Viega Viptool Master (Excel-Export)"
    accepts_extensions = (".xlsx", ".xls")

    # ----- detection -------------------------------------------------------

    def can_handle(self, filename: str, content_head: bytes) -> bool:
        lower = filename.lower()
        if not any(lower.endswith(ext) for ext in self.accepts_extensions):
            return False
        if "viptool" in lower:
            return True
        # Fallback: inspect cells A1..E1 of the first worksheet for a brand marker.
        # ``content_head`` may be truncated; we accept that — if the head isn't a
        # complete zip we won't claim the file and the GenericTableImporter takes over.
        return _first_cells_mention_brand(content_head)

    # ----- parse -----------------------------------------------------------

    def parse(
        self,
        filename: str,
        content: bytes,
        mapping: ColumnMapping | None = None,
    ) -> HeatingDesignImportPreview:
        warnings: list[str] = [
            "Viptool-Adapter ist ein Default-Stub. Nach Lieferung einer echten "
            "Viptool-Beispieldatei wird das Mapping verfeinert."
        ]

        try:
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:
            raise HeatingImporterError(
                f"Excel-Datei konnte nicht geoeffnet werden: {exc}"
            ) from exc

        try:
            ws = _pick_worksheet(wb)
            header_row, headers = _find_header_row(ws)
            by_field, labels, units = _build_column_mapping(headers)

            # Apply user-overrides from ColumnMapping (UI corrections).
            if mapping and mapping.circuit_columns:
                for canonical, header_name in mapping.circuit_columns.items():
                    if not header_name:
                        continue
                    for idx, raw in enumerate(headers):
                        if raw and _normalise(raw) == _normalise(header_name):
                            by_field[canonical] = idx
                            labels[canonical] = raw
                            units[canonical] = _detect_unit(raw)
                            break

            circuits: list[HeatingCircuitBase] = []
            position = 0
            data_row_count = 0
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                if not row or not any(cell not in (None, "") for cell in row):
                    continue
                data_row_count += 1
                circuit = self._row_to_circuit(row, by_field, units, position)
                if circuit is None:
                    continue
                circuits.append(circuit)
                position += 1

            if data_row_count == 0:
                raise HeatingImporterError(
                    "Worksheet enthaelt keine Datenzeilen unter der Kopfzeile."
                )

            design = self._infer_design(mapping)

            preview = HeatingDesignImportPreview(
                source=self.source_name,
                source_file=filename,
                design=design,
                circuits=circuits,
                warnings=warnings,
                detected_columns=labels,
            )
            return preview
        finally:
            wb.close()

    # ----- helpers ---------------------------------------------------------

    def _row_to_circuit(
        self,
        row: tuple,
        by_field: dict[str, int],
        units: dict[str, str | None],
        position: int,
    ) -> HeatingCircuitBase | None:
        def cell(field: str) -> object:
            idx = by_field.get(field)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        heat_load_raw = _coerce_float(cell("heat_load_w"))
        volume_raw = _coerce_float(cell("volume_flow_lph"))
        pressure_raw = _coerce_float(cell("pressure_drop_pa"))
        length_raw = _coerce_float(cell("pipe_length_m"))
        kv_raw = _coerce_float(cell("kv_value"))

        circuit = HeatingCircuitBase(
            position=position,
            strand=_coerce_str(cell("strand")),
            room=_coerce_str(cell("room")),
            floor=_coerce_str(cell("floor")),
            radiator_type=_coerce_str(cell("radiator_type")),
            heat_load_w=_to_watts(heat_load_raw, units.get("heat_load_w")) if heat_load_raw is not None else None,
            volume_flow_lph=_to_lph(volume_raw, units.get("volume_flow_lph")) if volume_raw is not None else None,
            pressure_drop_pa=_to_pascal(pressure_raw, units.get("pressure_drop_pa")) if pressure_raw is not None else None,
            pipe_length_m=_to_meters(length_raw, units.get("pipe_length_m")) if length_raw is not None else None,
            valve_type=_coerce_str(cell("valve_type")),
            valve_preset=_coerce_str(cell("valve_preset")),
            kv_value=kv_raw,
        )

        # Skip completely-empty rows that survived the early-empty check
        # because of trailing whitespace cells.
        if all(
            getattr(circuit, attr) in (None, "", 0, 0.0)
            for attr in (
                "strand",
                "room",
                "radiator_type",
                "heat_load_w",
                "volume_flow_lph",
            )
        ):
            return None
        return circuit

    def _infer_design(self, mapping: ColumnMapping | None) -> HeatingDesignBase:
        """Viptool-Exports enthalten typischerweise keine Anlagenkenndaten in den
        Kreis-Tabellen. Wenn der Benutzer im UI Overrides geliefert hat, uebernehmen
        wir sie; ansonsten bleiben die Felder ``None`` und der Generator weist sie
        spaeter als "Offene Punkte" aus.
        """
        overrides: dict[str, float | str | None] = {}
        if mapping and mapping.design_overrides:
            overrides = dict(mapping.design_overrides)

        def _val(key: str) -> float | str | None:
            return overrides.get(key)

        def _val_float(key: str) -> float | None:
            v = overrides.get(key)
            if v is None:
                return None
            if isinstance(v, (int, float)):
                return float(v)
            try:
                return float(str(v).replace(",", "."))
            except ValueError:
                return None

        return HeatingDesignBase(
            system_type=_val("system_type") if isinstance(_val("system_type"), str) else None,
            supply_temp_c=_val_float("supply_temp_c"),
            return_temp_c=_val_float("return_temp_c"),
            delta_t_k=_val_float("delta_t_k"),
            pump_head_pa=_val_float("pump_head_pa"),
            total_volume_flow_lph=_val_float("total_volume_flow_lph"),
            pump_model=_val("pump_model") if isinstance(_val("pump_model"), str) else None,
        )
