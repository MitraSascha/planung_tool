"""Generic Excel/CSV importer for external heating-design tables.

External architects/planners deliver pipe-network calculations in many
shapes: free Excel exports, CSV dumps with German headers, etc. This
adapter is the catch-all that
  1. sniffs the file format (xlsx / xls / csv),
  2. locates the header row even if the file starts with meta-data,
  3. fuzzy-matches each header to a canonical ``HeatingCircuit`` field,
  4. converts known units to the canonical SI representation,
  5. extracts plant-level metadata (system type, supply/return temperatures,
     pump head, ...) from the leading "Label: Value" lines,
  6. returns a :class:`HeatingDesignImportPreview` so the frontend can show
     a column-mapping UI before persisting anything.

The auto-detection result is exposed through ``preview.detected_columns``
(``{canonical_field: source_column}``) so the UI can pre-fill the mapping
form. Unmappable columns or unknown units land in ``preview.warnings``.
"""
from __future__ import annotations

import csv
import difflib
import io
import re
import unicodedata
from typing import Any, Iterable, Sequence

from app.models.heating import (
    HeatingCircuitBase,
    HeatingDesignBase,
    HeatingDesignImportPreview,
)
from app.services.heating_importers.base import (
    KNOWN_CIRCUIT_FIELDS,
    ColumnMapping,
    HeatingImporter,
    HeatingImporterError,
)


# ----------------------------------------------------------------------
# Header-Synonym-Tabelle. Reihenfolge wirkt sich auf Tie-Break aus,
# Treffer wird ueber den hoechsten difflib-Score gewaehlt.
# ----------------------------------------------------------------------

_FIELD_SYNONYMS: dict[str, tuple[str, ...]] = {
    "strand": ("strang", "strand", "vlstrang", "stranggruppe", "kreisnr", "kreisnummer"),
    "room": ("raum", "raume", "raumbezeichnung", "wohneinheit", "we"),
    "floor": ("etage", "etegae", "geschoss", "stockwerk", "stock", "ebene"),
    "radiator_type": (
        "heizkorper",
        "heizkoerper",
        "radiator",
        "heizflachentyp",
        "hktyp",
    ),
    "area_sqm": (
        "flachem2",
        "flachem",
        "wohnflache",
        "raumflache",
        "nutzflache",
        "grundflache",
        "flaeche",
    ),
    "heat_load_w": (
        "heizlast",
        "heizleistung",
        "normheizlast",
        "leistungw",
        "leistungkw",
        "watt",
    ),
    "volume_flow_lph": (
        "volumenstrom",
        "massenstrom",
        "durchfluss",
        "vstrom",
        "vdot",
    ),
    "pressure_drop_pa": (
        "druckverlust",
        "deltap",
        "druckverlustpa",
        "druckverlustmbar",
    ),
    "pipe_length_m": (
        "rohrlange",
        "rohrlaenge",
        "rohrlength",
    ),
    "valve_type": (
        "ventil",
        "ventiltyp",
        "thermostatventil",
        "ventilfabrikat",
        "armatur",
    ),
    "valve_preset": (
        "voreinstellung",
        "voreinstellwert",
        "voreinst",
        "einstellung",
        "stellwert",
    ),
    "kv_value": ("kvwert", "kvs", "kvvalue"),
    "notes": ("bemerkung", "notiz", "kommentar", "anmerkung", "comment"),
}


# Felder, die als "Pflicht-Signatur" einer Heizungs-/Strang-Tabelle gelten.
# Mindestens EINES davon muss gemappt sein, sonst handelt es sich nicht um
# eine Heizungsauslegung (z.B. Angebot, Stueckliste, Bauphysik-Tabelle).
_CORE_REQUIRED_FIELDS: tuple[str, ...] = ("heat_load_w", "volume_flow_lph")

# Mindest-Score fuer Auto-Mapping. 0.50 war zu freizuegig und liess "Menge"
# auf "Laenge", "Name" auf "Raum", "Einkaufspreis" auf "Strang" durchgehen.
_AUTO_MAPPING_MIN_SCORE: float = 0.78


# Felder, die als numerisch interpretiert werden sollen.
_NUMERIC_FIELDS = {
    "area_sqm",
    "heat_load_w",
    "volume_flow_lph",
    "pressure_drop_pa",
    "pipe_length_m",
    "kv_value",
}


# Plant-level meta-data labels (system_type, supply/return temp, etc.).
_DESIGN_LABEL_SYNONYMS: dict[str, tuple[str, ...]] = {
    "system_type": ("system", "systemtyp", "anlagenart", "anlage", "heizungstyp"),
    "supply_temp_c": ("vorlauf", "vorlauftemperatur", "tv", "vl"),
    "return_temp_c": ("rucklauf", "ruecklauf", "rucklauftemperatur", "tr", "rl"),
    "total_volume_flow_lph": ("gesamtvolumenstrom", "anlagenvolumenstrom", "vgesamt"),
    "pump_head_pa": ("forderhohe", "foerderhoehe", "pumpenforderhohe", "pumpendruck"),
    "pump_model": ("pumpe", "pumpenmodell", "umwalzpumpe", "umwalzungspumpe"),
}


# ----------------------------------------------------------------------
# Public helpers (also covered by unit tests).
# ----------------------------------------------------------------------


def normalize_header(header: str) -> str:
    """Lowercase, strip, remove unit suffix in brackets, replace umlauts.

    >>> normalize_header("Heizlast [W]")
    'heizlast'
    >>> normalize_header("Vorlauf-Temperatur (°C)")
    'vorlauftemperatur'
    """
    if header is None:
        return ""
    text = str(header).strip().lower()
    # Strip bracketed unit suffixes like "[W]", "(m)", "{kg/h}".
    text = re.sub(r"[\[\(\{][^\]\)\}]*[\]\)\}]", " ", text)
    # Strip ASCII control / non-printable characters.
    text = "".join(ch for ch in text if ch.isprintable())
    # Replace German umlauts and eszett.
    replacements = {"ä": "a", "ö": "o", "ü": "u", "ß": "ss"}
    for src, dst in replacements.items():
        text = text.replace(src, dst)
    # Strip remaining accents via NFKD decomposition.
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    # Reduce all non-alphanumeric runs to a single space, then collapse.
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", "", text).strip()


_UNIT_HINT_RE = re.compile(
    r"\b("
    r"k?w|mw|watt|"
    r"kw?h|"
    r"l\s*/\s*h|l/?s|"
    r"m3\s*/\s*h|m\^?3/h|m³\s*/\s*h|"
    r"kg\s*/\s*h|"
    r"k?pa|pascal|bar|m?bar|"
    r"k?m|cm|mm|meter|"
    r"°\s*c|grad"
    r")\b",
    flags=re.IGNORECASE,
)


def _extract_unit(header: str) -> str | None:
    """Return the unit token (case-preserved) found in the header.

    Looks for ``[unit]`` / ``(unit)`` / ``{unit}`` first, then for a unit-like
    token after a newline (Excel headers often place units on a second line,
    e.g. ``"Heizlast\\nQ / kW"`` or ``"Volumen-\\nstrom\\nV̇ m³/h"``),
    and finally for an inline unit token.
    """
    if header is None:
        return None
    text = str(header)
    match = re.search(r"[\[\(\{]([^\]\)\}]+)[\]\)\}]", text)
    if match:
        unit = match.group(1).strip()
        if unit:
            return unit
    # Inspect lines from bottom up — units typically sit on the last line.
    for line in reversed([ln.strip() for ln in text.splitlines() if ln.strip()]):
        hit = _UNIT_HINT_RE.search(line)
        if hit:
            return hit.group(1)
    return None


def fuzzy_match_canonical(header: str) -> tuple[str | None, float]:
    """Return ``(canonical_field, score)`` for the best fuzzy match.

    ``score`` is the difflib SequenceMatcher ratio against the closest
    synonym. ``canonical_field`` is None when no synonym scores >= 0.5.
    """
    normalized = normalize_header(header)
    if not normalized:
        return None, 0.0
    best_field: str | None = None
    best_score = 0.0
    for canonical, synonyms in _FIELD_SYNONYMS.items():
        for synonym in synonyms:
            if not synonym:
                continue
            # Exact match wins outright.
            if synonym == normalized:
                score = 1.0
            # Containment is a strong signal — but only for tokens long
            # enough to be discriminating (>= 4 chars). Short tokens like
            # "q" / "w" / "m" trigger far too many false positives.
            elif len(synonym) >= 4 and synonym in normalized:
                score = max(0.85, len(synonym) / max(len(normalized), 1))
            else:
                score = difflib.SequenceMatcher(None, normalized, synonym).ratio()
            if score > best_score:
                best_score = score
                best_field = canonical
    if best_score < 0.5:
        return None, best_score
    return best_field, best_score


# ----------------------------------------------------------------------
# Unit conversion.
# ----------------------------------------------------------------------


def _convert_value(field: str, value: Any, unit: str | None) -> tuple[Any, str | None]:
    """Convert ``value`` to the canonical unit for ``field``.

    Returns ``(converted_value, warning_or_None)``.  Warnings are emitted
    when a unit was given but not recognised — we keep the original
    value to avoid silently corrupting numbers.
    """
    if value is None or value == "":
        return None, None

    if field not in _NUMERIC_FIELDS:
        # Keep text fields as-is (radiator_type, valve_type, valve_preset,
        # notes, strand, room, floor).
        return str(value).strip(), None

    # Parse numeric (tolerate German decimal comma + thousand separators).
    try:
        if isinstance(value, (int, float)):
            num = float(value)
        else:
            cleaned = str(value).strip().replace(" ", "")
            if not cleaned:
                return None, None
            # If both ',' and '.' appear, the rightmost one is the decimal
            # separator (typical German "1.234,56" or English "1,234.56").
            if "," in cleaned and "." in cleaned:
                if cleaned.rfind(",") > cleaned.rfind("."):
                    cleaned = cleaned.replace(".", "").replace(",", ".")
                else:
                    cleaned = cleaned.replace(",", "")
            else:
                cleaned = cleaned.replace(",", ".")
            num = float(cleaned)
    except (TypeError, ValueError):
        return None, f"Wert nicht numerisch fuer Feld '{field}': {value!r}"

    if unit is None:
        return num, None

    unit_norm = unit.strip().lower().replace(" ", "")
    unit_norm = unit_norm.replace("²", "2").replace("³", "3")

    warning: str | None = None
    if field == "heat_load_w":
        if unit_norm in ("w", "watt"):
            pass
        elif unit_norm in ("kw", "kilowatt"):
            num *= 1000.0
        elif unit_norm in ("mw", "megawatt"):
            num *= 1_000_000.0
        else:
            warning = f"Unbekannte Einheit fuer heat_load_w: {unit!r}"
    elif field == "volume_flow_lph":
        if unit_norm in ("l/h", "lh", "liter/h"):
            pass
        elif unit_norm in ("kg/h", "kgh"):
            # 1 kg Wasser ~ 1 l (vereinfacht).
            pass
        elif unit_norm in ("m3/h", "m^3/h", "m3h"):
            num *= 1000.0
        elif unit_norm in ("l/s", "ls"):
            num *= 3600.0
        else:
            warning = f"Unbekannte Einheit fuer volume_flow_lph: {unit!r}"
    elif field == "pressure_drop_pa":
        if unit_norm in ("pa", "pascal"):
            pass
        elif unit_norm in ("hpa", "mbar"):
            num *= 100.0
        elif unit_norm in ("kpa",):
            num *= 1000.0
        elif unit_norm in ("bar",):
            num *= 100_000.0
        else:
            warning = f"Unbekannte Einheit fuer pressure_drop_pa: {unit!r}"
    elif field == "pipe_length_m":
        if unit_norm in ("m", "meter"):
            pass
        elif unit_norm in ("cm",):
            num /= 100.0
        elif unit_norm in ("mm",):
            num /= 1000.0
        elif unit_norm in ("km",):
            num *= 1000.0
        else:
            warning = f"Unbekannte Einheit fuer pipe_length_m: {unit!r}"
    # kv_value is dimensionless, no conversion.

    return num, warning


# ----------------------------------------------------------------------
# File-format handling.
# ----------------------------------------------------------------------


def _decode_csv_bytes(content: bytes) -> str:
    """Decode CSV bytes, trying UTF-8 (with BOM) then latin-1."""
    if content[:3] == b"\xef\xbb\xbf":
        return content[3:].decode("utf-8", errors="replace")
    for encoding in ("utf-8", "cp1252", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def _sniff_csv(text: str) -> csv.Dialect:
    """Detect delimiter; fall back to comma."""
    sample = text[:4096]
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        # Most German exports use ';'. If we see more ';' than ',' fall back to that.
        first_lines = sample.splitlines()[:5]
        semi = sum(line.count(";") for line in first_lines)
        comma = sum(line.count(",") for line in first_lines)
        tab = sum(line.count("\t") for line in first_lines)
        if tab >= semi and tab >= comma and tab > 0:
            return csv.excel_tab  # type: ignore[return-value]
        if semi > comma:
            class _Semicolon(csv.excel):
                delimiter = ";"

            return _Semicolon()  # type: ignore[return-value]
        return csv.excel()  # type: ignore[return-value]


def _read_csv_rows(content: bytes) -> list[list[str]]:
    text = _decode_csv_bytes(content)
    if not text.strip():
        return []
    dialect = _sniff_csv(text)
    reader = csv.reader(io.StringIO(text), dialect=dialect)
    rows: list[list[str]] = []
    for raw_row in reader:
        # Trim trailing all-empty cells.
        row = [cell.strip() if isinstance(cell, str) else cell for cell in raw_row]
        while row and (row[-1] is None or row[-1] == ""):
            row.pop()
        rows.append(row)
    return rows


def _read_xlsx_sheets(content: bytes) -> dict[str, list[list[Any]]]:
    """Read every sheet of an XLSX. Returns ``{sheet_name: rows}``."""
    try:
        import openpyxl  # type: ignore
    except ImportError as exc:  # pragma: no cover - requirements.txt pins it
        raise HeatingImporterError(
            "openpyxl ist nicht installiert; .xlsx kann nicht gelesen werden"
        ) from exc

    try:
        workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        raise HeatingImporterError(f"Excel-Datei konnte nicht geoeffnet werden: {exc}") from exc

    result: dict[str, list[list[Any]]] = {}
    for name in workbook.sheetnames:
        sheet = workbook[name]
        rows: list[list[Any]] = []
        for raw_row in sheet.iter_rows(values_only=True):
            row = list(raw_row)
            while row and (row[-1] is None or row[-1] == ""):
                row.pop()
            rows.append(row)
        result[name] = rows
    workbook.close()
    return result


def _read_xlsx_rows(content: bytes) -> list[list[Any]]:
    """Backwards-compatible single-sheet read (first sheet)."""
    sheets = _read_xlsx_sheets(content)
    if not sheets:
        return []
    first_name = next(iter(sheets))
    return sheets[first_name]


def _read_xls_rows(content: bytes) -> list[list[Any]]:
    try:
        import xlrd  # type: ignore
    except ImportError as exc:  # pragma: no cover
        raise HeatingImporterError(
            "xlrd ist nicht installiert; .xls kann nicht gelesen werden"
        ) from exc

    try:
        book = xlrd.open_workbook(file_contents=content)
    except Exception as exc:
        raise HeatingImporterError(f".xls-Datei konnte nicht geoeffnet werden: {exc}") from exc

    sheet = book.sheet_by_index(0)
    rows: list[list[Any]] = []
    for row_idx in range(sheet.nrows):
        raw = sheet.row_values(row_idx)
        row = list(raw)
        while row and (row[-1] is None or row[-1] == ""):
            row.pop()
        rows.append(row)
    return rows


# ----------------------------------------------------------------------
# Header detection.
# ----------------------------------------------------------------------


def _row_is_plausible_header(row: Sequence[Any]) -> tuple[bool, int]:
    """Return ``(is_header, hit_count)`` for a candidate row.

    A row counts as a header when at least 3 of its non-empty cells map
    to a canonical heating field with score >= 0.5. ``hit_count`` is the
    total number of matched cells (including weak fuzzy matches).
    """
    hits, _strong = _score_header_row(row)
    non_empty = sum(
        1 for cell in row if cell is not None and str(cell).strip() != ""
    )
    return (hits >= 3 and non_empty >= 3), hits


def _score_header_row(row: Sequence[Any]) -> tuple[int, int]:
    """Return ``(weak_or_better_hits, strong_hits)`` for a candidate row.

    A *strong* hit is an exact synonym match or a containment match where
    the synonym is fully present in the normalised header (score >= 0.85).
    Strong hits are reliable signals; weak fuzzy hits (0.5–0.85) often
    misfire on free-form parameter sheets and should not dominate sheet
    selection on their own.
    """
    best_per_field = _row_best_per_field(row)
    weak = len(best_per_field)
    strong = sum(1 for s in best_per_field.values() if s >= 0.85)
    return weak, strong


# Two canonical fields together are the signature of a heating-circuit
# overview / strand calculation. Sheets that map both strongly outrank
# building-physics or parameter sheets that only carry one of them.
_CORE_HEATING_FIELDS: tuple[str, str] = ("heat_load_w", "volume_flow_lph")


def _row_best_per_field(row: Sequence[Any]) -> dict[str, float]:
    """Best fuzzy score per canonical field for one row.

    Tracks the BEST score per field so a weak false-positive earlier in
    the row doesn't crowd out a strong proper match later (e.g. "WE"
    weakly hitting heat_load_w before "Heizlast" strongly does).
    """
    best: dict[str, float] = {}
    for cell in row:
        if cell is None:
            continue
        text = str(cell).strip()
        if not text:
            continue
        canonical, score = fuzzy_match_canonical(text)
        if canonical is None or score < 0.5:
            continue
        if score > best.get(canonical, 0.0):
            best[canonical] = score
    return best


def _find_header_row(rows: Sequence[Sequence[Any]]) -> int | None:
    best_idx: int | None = None
    best_hits = 0
    # Try the first 20 rows only — beyond that it's unlikely.
    for idx, row in enumerate(rows[:20]):
        is_header, hits = _row_is_plausible_header(row)
        if is_header and hits > best_hits:
            best_hits = hits
            best_idx = idx
    return best_idx


# ----------------------------------------------------------------------
# Plant metadata extraction.
# ----------------------------------------------------------------------


def _parse_design_meta(meta_rows: Sequence[Sequence[Any]]) -> HeatingDesignBase:
    """Look for "Label: Value" style cells in pre-header rows."""
    system_type: str | None = None
    supply_temp_c: float | None = None
    return_temp_c: float | None = None
    total_volume_flow_lph: float | None = None
    pump_head_pa: float | None = None
    pump_model: str | None = None

    def _try_set(label: str, value: Any) -> None:
        nonlocal system_type, supply_temp_c, return_temp_c
        nonlocal total_volume_flow_lph, pump_head_pa, pump_model
        normalized_label = normalize_header(label)
        if not normalized_label:
            return
        for canonical, synonyms in _DESIGN_LABEL_SYNONYMS.items():
            if any(syn in normalized_label for syn in synonyms):
                unit = _extract_unit(label)
                if canonical == "system_type":
                    system_type = str(value).strip() or None
                elif canonical == "pump_model":
                    pump_model = str(value).strip() or None
                elif canonical == "supply_temp_c":
                    converted, _ = _convert_value("pipe_length_m", value, unit)
                    # Temperatures aren't in _NUMERIC_FIELDS — parse directly.
                    supply_temp_c = _to_float(value)
                elif canonical == "return_temp_c":
                    return_temp_c = _to_float(value)
                elif canonical == "total_volume_flow_lph":
                    num, _ = _convert_value("volume_flow_lph", value, unit)
                    if isinstance(num, (int, float)):
                        total_volume_flow_lph = float(num)
                elif canonical == "pump_head_pa":
                    num, _ = _convert_value("pressure_drop_pa", value, unit)
                    if isinstance(num, (int, float)):
                        pump_head_pa = float(num)
                return

    for row in meta_rows:
        cells = [c for c in row if c is not None and str(c).strip() != ""]
        # Pattern A: label-value pairs in adjacent cells.
        for idx in range(len(cells) - 1):
            label_cell = str(cells[idx]).strip()
            if label_cell.endswith(":") or label_cell.endswith("="):
                _try_set(label_cell.rstrip(":=").strip(), cells[idx + 1])
        # Pattern B: single-cell "Label: Value".
        for cell in cells:
            text = str(cell)
            if ":" in text:
                label, _, value = text.partition(":")
                if label.strip() and value.strip():
                    _try_set(label, value.strip())

    delta_t_k: float | None = None
    if supply_temp_c is not None and return_temp_c is not None:
        delta_t_k = supply_temp_c - return_temp_c

    return HeatingDesignBase(
        system_type=system_type,
        supply_temp_c=supply_temp_c,
        return_temp_c=return_temp_c,
        delta_t_k=delta_t_k,
        pump_head_pa=pump_head_pa,
        total_volume_flow_lph=total_volume_flow_lph,
        pump_model=pump_model,
        notes=None,
    )


def _to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    # Take the first numeric token (handles "55 C", "55°C", "55.0 Grad").
    match = re.match(r"-?\d+(?:[\.,]\d+)?", text)
    if not match:
        return None
    try:
        return float(match.group(0).replace(",", "."))
    except (TypeError, ValueError):
        return None


# ----------------------------------------------------------------------
# The importer itself.
# ----------------------------------------------------------------------


class GenericTableImporter(HeatingImporter):
    """Catch-all importer for free-form Excel/CSV heating tables."""

    source_name = "generic_table"
    display_name = "Generic Excel/CSV (extern)"
    accepts_extensions = (".xlsx", ".xls", ".csv")

    def can_handle(self, filename: str, content_head: bytes) -> bool:
        lower = filename.lower()
        return any(lower.endswith(ext) for ext in self.accepts_extensions)

    # ------------------------------------------------------------------
    # parse()
    # ------------------------------------------------------------------

    def parse(
        self,
        filename: str,
        content: bytes,
        mapping: ColumnMapping | None = None,
    ) -> HeatingDesignImportPreview:
        if not content:
            raise HeatingImporterError("Datei ist leer.")

        lower = filename.lower()
        if lower.endswith(".xlsx"):
            sheets = _read_xlsx_sheets(content)
            if not sheets:
                raise HeatingImporterError("XLSX-Datei enthaelt keine Sheets.")
            return self._parse_multi_sheet(filename, sheets, mapping)

        # CSV / .xls — single-table flow
        if lower.endswith(".xls"):
            rows = _read_xls_rows(content)
        elif lower.endswith(".csv"):
            rows = _read_csv_rows(content)
        else:
            rows = _read_csv_rows(content)
        return self._parse_single_sheet(filename, None, rows, mapping, [])

    # ------------------------------------------------------------------
    # Sheet classification (which kind of data is on this sheet?)
    # ------------------------------------------------------------------

    @staticmethod
    def _classify_sheet(rows: list[list[Any]]) -> str:
        """Return one of: 'aggregate' | 'bauphysik' | 'schema' | 'parameter'.

        - 'aggregate'  — Wohnungs-Aggregat (WE / Etage / Heizlast / Volumenstrom)
                         — das ist, was unser HeatingDesign-Modell abbildet.
        - 'bauphysik'  — DIN-EN 12831 Bauteil-Tabelle (Bauteil / U-Wert / Fläche / ΔT)
                         — separater Importer (in Planung).
        - 'schema'     — Visuelles Strang-Schema (keine klassische Tabellenstruktur).
        - 'parameter'  — Konstanten-Liste, kein Datentabellen-Sheet.
        """
        # Strang-Schema-Marker in den ersten 5 Zeilen
        for row in rows[:5]:
            joined = " ".join(str(c) for c in row if c is not None).lower()
            if any(
                marker in joined
                for marker in ("strangschema", "starngschema", "rohrmaterial", "δp spreizung")
            ):
                return "schema"

        # Bauphysik-Marker: Wenn ein Header-Row die typischen Bauphysik-Spalten
        # zusammen enthält, ist es DIN-EN 12831 Detail-Berechnung.
        bauphysik_signals = {"bauteil", "uwert", "uvalue", "wbruecken", "flachem2", "deltatk", "fx"}
        for row in rows[:5]:
            normalized = {normalize_header(c) for c in row if c is not None and str(c).strip()}
            hits = bauphysik_signals & normalized
            # "bauteil" alleine reicht schon — sehr eindeutiger Marker fuer DIN-EN.
            if "bauteil" in hits or len(hits) >= 2:
                return "bauphysik"

        # Aggregat-Heuristik: enthaelt heat_load und/oder volume_flow strong
        # ABER NICHT die Bauphysik-Marker → klassische Strang-/Aggregat-Tabelle.
        for row in rows[:20]:
            per = _row_best_per_field(row)
            strong_core = sum(
                1 for f in _CORE_HEATING_FIELDS if per.get(f, 0.0) >= 0.85
            )
            if strong_core >= 1:
                return "aggregate"

        return "parameter"

    # ------------------------------------------------------------------
    # Multi-sheet XLSX flow
    # ------------------------------------------------------------------

    _SKIP_REASONS = {
        "bauphysik": (
            "Bauphysik-Berechnung (DIN-EN 12831, Bauteilebene) erkannt — "
            "hier nicht nutzbar. Eigener Importer 'Bauphysik / "
            "Sanierungs-Analyse' kommt."
        ),
        "schema": (
            "Visuelles Strang-Schema erkannt — nicht maschinenlesbar. "
            "Bitte manuell oder aus Strang-Tool-Export (VIPtool, ETU, …) importieren."
        ),
        "parameter": (
            "Parameter-/Konstanten-Liste — keine Datentabelle."
        ),
    }

    @staticmethod
    def _extract_strand_assignments(
        sheets: dict[str, list[list[Any]]],
    ) -> dict[str, list[str]]:
        """Scan strang-schema sheets for 'WE X' / 'Wohnung X' mentions.

        Sheets named 'Strang N' (or 'Kellerleitung') are visual diagrams,
        not tabular data — but they DO contain WE references that tell us
        which Wohnung belongs to which strand. We harvest that info so the
        aggregate-data rows ("Wohnung 14") can be auto-assigned a strand.
        """
        assignments: dict[str, set[str]] = {}
        strand_name_re = re.compile(r"^(strang)\s*(\d+)\b", re.IGNORECASE)
        we_re = re.compile(r"^(?:WE|Wohnung)\s*(\d+)\s*$", re.IGNORECASE)
        for sheet_name, rows in sheets.items():
            m = strand_name_re.match(sheet_name.strip())
            if not m:
                continue
            strand_no = m.group(2)
            for row in rows:
                for cell in row:
                    if cell is None:
                        continue
                    text = str(cell).strip()
                    if not text:
                        continue
                    we_match = we_re.match(text)
                    if we_match:
                        we_key = f"Wohnung {we_match.group(1)}"
                        assignments.setdefault(we_key, set()).add(strand_no)
        # Sort strand numbers numerically per WE.
        return {
            we: sorted(strands, key=lambda s: int(s) if s.isdigit() else 999)
            for we, strands in assignments.items()
        }

    def _parse_multi_sheet(
        self,
        filename: str,
        sheets: dict[str, list[list[Any]]],
        mapping: ColumnMapping | None,
    ) -> HeatingDesignImportPreview:
        """Process every sheet that looks like a heating table and merge.

        - Skips sheets without a plausible heating-table header (e.g.
          a cover page, a parameter sheet, a legend) with a warning.
        - Merges all valid sheets' circuits into one design, with a
          continuous position index across sheets.
        - Each circuit's ``notes`` is prefixed with ``[Sheet: <name>]`` so
          users can trace a row back to its source tab.
        - The first valid sheet contributes plant-level metadata
          (system_type, supply/return temp, etc.); later sheets are
          consulted only for fields the first sheet didn't fill.
        """
        all_circuits: list[HeatingCircuitBase] = []
        all_warnings: list[str] = []
        merged_design: HeatingDesignBase | None = None
        merged_detected: dict[str, str] = {}
        merged_source_columns: dict[str, list[str]] = {}
        used_sheets: list[tuple[str, int]] = []
        skipped_sheets: list[tuple[str, str]] = []
        position_offset = 0

        for sheet_name, raw_rows in sheets.items():
            rows = [
                r for r in raw_rows
                if any(c is not None and str(c).strip() != "" for c in r)
            ]
            if not rows:
                skipped_sheets.append((sheet_name, "leer"))
                continue

            kind = self._classify_sheet(rows)
            if kind != "aggregate":
                skipped_sheets.append((sheet_name, self._SKIP_REASONS.get(kind, kind)))
                continue

            try:
                sub_preview = self._parse_single_sheet(
                    filename, sheet_name, rows, mapping, []
                )
            except HeatingImporterError:
                # Long error messages don't help here — they're already
                # grouped with other rejects under a short reason.
                skipped_sheets.append(
                    (sheet_name, "kein nutzbarer Heizlast-Datenblock")
                )
                continue

            sheet_circuits = sub_preview.circuits
            if not sheet_circuits:
                skipped_sheets.append((sheet_name, "keine Datenzeilen"))
                continue

            # Re-index positions continuous across sheets, prefix notes
            # with origin sheet name so users can trace rows back.
            renumbered: list[HeatingCircuitBase] = []
            for offset, circuit in enumerate(sheet_circuits):
                payload = circuit.model_dump()
                payload["position"] = position_offset + offset
                origin = f"[Sheet: {sheet_name}]"
                existing_notes = payload.get("notes")
                payload["notes"] = (
                    f"{origin} {existing_notes}".strip()
                    if existing_notes
                    else origin
                )
                renumbered.append(HeatingCircuitBase(**payload))
            all_circuits.extend(renumbered)
            position_offset += len(renumbered)

            used_sheets.append((sheet_name, len(renumbered)))
            for w in sub_preview.warnings:
                tagged = f"[{sheet_name}] {w}"
                if tagged not in all_warnings:
                    all_warnings.append(tagged)
            if merged_design is None:
                merged_design = sub_preview.design
                merged_detected = dict(sub_preview.detected_columns)
                merged_source_columns = dict(sub_preview.source_columns)
            else:
                # Fill missing design fields from later sheets.
                base = merged_design.model_dump()
                for key, val in sub_preview.design.model_dump().items():
                    if base.get(key) is None and val is not None:
                        base[key] = val
                merged_design = HeatingDesignBase(**base)
                # Track new columns we've seen but don't overwrite earlier picks.
                for key, val in sub_preview.detected_columns.items():
                    merged_detected.setdefault(key, val)
                for header, samples in sub_preview.source_columns.items():
                    merged_source_columns.setdefault(header, samples)

        # Harvest WE→Strang from any Strang-schema sheets in the workbook
        # and apply to circuits whose room is 'Wohnung X'.
        strand_map = self._extract_strand_assignments(sheets)
        if strand_map and all_circuits:
            applied = 0
            for circuit in all_circuits:
                room = (circuit.room or "").strip()
                if room in strand_map and not circuit.strand:
                    circuit.strand = ", ".join(strand_map[room])
                    applied += 1
            if applied:
                all_warnings.insert(
                    1,
                    f"Strang-Zuordnung aus Schema-Sheets: {applied} Wohnungen "
                    f"automatisch zugeordnet (Quelle: {', '.join(sn for sn in sheets if sn.lower().startswith('strang'))}).",
                )

        if not used_sheets:
            details = "; ".join(f"{n}: {r}" for n, r in skipped_sheets) or "keine"
            raise HeatingImporterError(
                "Keines der Sheets enthielt eine erkennbare Heizungstabelle. "
                f"Sheets geprueft: {details}"
            )

        # Top-of-list summary so the user sees the punchline first.
        summary = "Verwendet: " + ", ".join(f"{n} ({c} Pos)" for n, c in used_sheets)
        all_warnings.insert(0, summary)

        # Group skipped sheets by reason so the panel stays readable.
        # 12 separate "Sheet X übersprungen: …" lines = noise; 3 grouped
        # lines = scannable.
        by_reason: dict[str, list[str]] = {}
        for sheet_name, reason in skipped_sheets:
            short = reason.split("—")[0].strip().rstrip(":").strip()
            # Strip trailing punctuation/whitespace; coalesce duplicates.
            by_reason.setdefault(short, []).append(sheet_name)
        for reason, sheet_names in by_reason.items():
            joined = ", ".join(f"'{n}'" for n in sheet_names)
            all_warnings.append(f"Sheets übersprungen ({reason}): {joined}")

        if merged_design is None:
            merged_design = HeatingDesignBase()

        return HeatingDesignImportPreview(
            source=self.source_name,
            source_file=filename,
            design=merged_design,
            circuits=all_circuits,
            warnings=all_warnings,
            detected_columns=merged_detected,
            source_columns=merged_source_columns,
        )

    # ------------------------------------------------------------------
    # Single-sheet flow (called per CSV, per .xls, or per XLSX sheet)
    # ------------------------------------------------------------------

    def _parse_single_sheet(
        self,
        filename: str,
        sheet_name: str | None,
        rows: list[list[Any]],
        mapping: ColumnMapping | None,
        warnings_prelude: list[str],
    ) -> HeatingDesignImportPreview:
        rows = [row for row in rows if any(c is not None and str(c).strip() != "" for c in row)]
        if not rows:
            raise HeatingImporterError("Datei enthaelt keine lesbaren Zeilen.")

        header_idx = _find_header_row(rows)
        if header_idx is None:
            hint = (
                f" Sheet '{sheet_name}' enthaelt keine tabellarische Struktur."
                if sheet_name
                else ""
            )
            raise HeatingImporterError(
                "Keine plausible Header-Zeile gefunden. Mindestens drei "
                "Spalten muessen auf Standardfelder (Heizlast, Volumenstrom, "
                "Druckverlust, ...) gemappt werden koennen." + hint
            )

        meta_rows = rows[:header_idx]
        header_row = rows[header_idx]
        data_rows = rows[header_idx + 1:]

        headers: list[str] = [str(c).strip() if c is not None else "" for c in header_row]
        warnings: list[str] = list(warnings_prelude)

        detected_columns: dict[str, str] = self._auto_detect_columns(headers, warnings)

        manual_override = mapping is not None and bool(mapping.circuit_columns)
        if manual_override:
            valid: dict[str, str] = {}
            header_lookup = {h: h for h in headers}
            for canonical, source_col in mapping.circuit_columns.items():
                if source_col in header_lookup:
                    valid[canonical] = source_col
                else:
                    warnings.append(
                        f"Manuelles Mapping: Spalte '{source_col}' nicht in Datei gefunden"
                    )
            detected_columns = valid

        if not manual_override:
            detected_columns = self._validate_numeric_mappings(
                detected_columns, headers, data_rows, warnings
            )

        if not manual_override and not any(
            f in detected_columns for f in _CORE_REQUIRED_FIELDS
        ):
            raise HeatingImporterError(
                "Diese Datei enthaelt keine erkennbare Heizungsauslegung: "
                "Es konnte weder eine Spalte fuer 'Heizlast (W)' noch fuer "
                "'Volumenstrom (l/h)' gefunden werden. "
                "Bitte pruefen Sie, ob Sie die richtige Datei "
                "(Strangberechnung bzw. Heizlast-Berechnung) hochgeladen "
                "haben. Erkannte Spalten: "
                + (", ".join(headers) if headers else "(keine)")
            )

        circuits: list[HeatingCircuitBase] = []
        unknown_units_seen: set[str] = set()
        canonical_index: list[tuple[str, int, str | None]] = []
        for canonical, source_col in detected_columns.items():
            try:
                idx = headers.index(source_col)
            except ValueError:
                continue
            canonical_index.append((canonical, idx, _extract_unit(source_col)))

        # Stop reading once we've seen real data and then run into a
        # gap — many heating-XLSX sheets append parameter blocks or
        # totals after the actual circuit rows.
        consecutive_invalid = 0
        # Row labels that indicate a summary / total / footer row — these
        # are aggregates of the rows above and must not be counted as
        # individual data points (otherwise totals double).
        _summary_labels = {
            "gesamt", "summe", "total", "summen", "sum", "gesamtsumme",
            "zwischensumme", "endsumme",
        }

        for position, raw_row in enumerate(data_rows):
            row_is_blank = not any(c is not None and str(c).strip() != "" for c in raw_row)
            if row_is_blank:
                consecutive_invalid += 1
                if circuits and consecutive_invalid >= 2:
                    break
                continue
            field_values: dict[str, Any] = {"position": position}
            had_numeric_core = False
            for canonical, idx, unit in canonical_index:
                raw_value = raw_row[idx] if idx < len(raw_row) else None
                converted, warn = _convert_value(canonical, raw_value, unit)
                if warn and warn not in unknown_units_seen:
                    warnings.append(warn)
                    unknown_units_seen.add(warn)
                if converted is not None:
                    field_values[canonical] = converted
                    if canonical in _CORE_HEATING_FIELDS and isinstance(converted, (int, float)):
                        had_numeric_core = True
            non_position = {k: v for k, v in field_values.items() if k != "position"}
            if not non_position:
                consecutive_invalid += 1
                if circuits and consecutive_invalid >= 2:
                    break
                continue
            # Skip summary/totals rows (they aggregate the rows above and
            # would otherwise double-count when summed).
            room_label = str(field_values.get("room", "")).strip().lower()
            if room_label in _summary_labels:
                # Don't reset consecutive_invalid — treat as a stop signal.
                break
            # Once we've started collecting circuits, abort if a row has
            # no core (heat_load_w / volume_flow_lph) value — that's
            # almost always the start of a parameter block ("Vorlauf",
            # "Rohrtabelle", …).
            if circuits and not had_numeric_core:
                consecutive_invalid += 1
                if consecutive_invalid >= 2:
                    break
                continue
            consecutive_invalid = 0
            try:
                circuits.append(HeatingCircuitBase(**field_values))
            except Exception as exc:  # pragma: no cover - pydantic is permissive
                warnings.append(f"Zeile {position + 1} uebersprungen: {exc}")

        if not circuits:
            raise HeatingImporterError(
                "Keine Datenzeilen erkannt — moeglicherweise sind alle Werte leer "
                "oder die erkannten Spalten passen nicht zu numerischen Feldern."
            )

        # Plant metadata extracted from rows above the header.
        design = _parse_design_meta(meta_rows)

        # Optional: apply design_overrides from the explicit mapping.
        if mapping is not None and mapping.design_overrides:
            design_dict = design.model_dump()
            for key, value in mapping.design_overrides.items():
                if key in design_dict:
                    design_dict[key] = value
            # Recompute delta_t_k after overrides if both temps are known.
            sup = design_dict.get("supply_temp_c")
            ret = design_dict.get("return_temp_c")
            if isinstance(sup, (int, float)) and isinstance(ret, (int, float)):
                design_dict["delta_t_k"] = float(sup) - float(ret)
            design = HeatingDesignBase(**design_dict)

        # Collect up to 5 sample values per source column so the UI can show
        # "WE (Wohnung 1, Wohnung 2, ...)" in the manual-mapping dropdown.
        source_columns: dict[str, list[str]] = {}
        for col_idx, header in enumerate(headers):
            if not header:
                continue
            samples: list[str] = []
            for raw_row in data_rows:
                if col_idx >= len(raw_row):
                    continue
                v = raw_row[col_idx]
                if v is None:
                    continue
                text = str(v).strip()
                if not text:
                    continue
                if text in samples:
                    continue
                samples.append(text[:50])
                if len(samples) >= 5:
                    break
            source_columns[header] = samples

        # Surface unmapped headers as one summary line instead of one warning
        # per column. Skip columns with no data rows (typically title /
        # metadata cells like 'BV Mareschstraße 1 ...').
        mapped_sources = set(detected_columns.values())
        leftovers = [
            h for h in headers
            if h and h not in mapped_sources and source_columns.get(h)
        ]
        if leftovers:
            warnings.append(
                "Nicht ins Heizlast-Modell übernommen: " + ", ".join(leftovers)
            )

        return HeatingDesignImportPreview(
            source=self.source_name,
            source_file=filename,
            design=design,
            circuits=circuits,
            warnings=warnings,
            detected_columns=detected_columns,
            source_columns=source_columns,
        )

    # ------------------------------------------------------------------
    # internals
    # ------------------------------------------------------------------

    @staticmethod
    def _pick_best_sheet(
        sheets: dict[str, list[list[Any]]],
        warnings: list[str],
    ) -> tuple[str | None, list[list[Any]]]:
        """Choose the sheet with the most plausible header row.

        XLSX workbooks often include summary / overview sheets, per-floor
        calculations, and per-strand schemas alongside free-form parameter
        sheets. The first sheet is rarely the tabular one. Score each sheet
        by its best header-row hit count and pick the winner.
        """
        if not sheets:
            return None, []

        best_name: str | None = None
        best_rows: list[list[Any]] = []
        best_key: tuple[int, int, int] = (0, 0, 0)
        for name, raw_rows in sheets.items():
            non_empty = [
                r for r in raw_rows
                if any(c is not None and str(c).strip() != "" for c in r)
            ]
            if not non_empty:
                continue
            best_weak = 0
            best_strong = 0
            best_core = 0
            for row in non_empty[:20]:
                fields = _row_best_per_field(row)
                weak = len(fields)
                strong = sum(1 for s in fields.values() if s >= 0.85)
                core = sum(
                    1 for f in _CORE_HEATING_FIELDS
                    if fields.get(f, 0.0) >= 0.85
                )
                if (core, strong, weak) > (best_core, best_strong, best_weak):
                    best_core = core
                    best_strong = strong
                    best_weak = weak
            # Rank sheets by:
            #  1. how many CORE heating fields (Heizlast + Volumenstrom)
            #     map strongly — that's the signature of a circuit table;
            #  2. total strong matches;
            #  3. total weak-or-better matches.
            # Without (1) the picker prefers free-form building-physics
            # sheets that accidentally trigger many low-confidence
            # fuzzy matches over a clean summary table.
            key = (best_core, best_strong, best_weak)
            if key > best_key:
                best_key = key
                best_name = name
                best_rows = raw_rows

        if best_name is None:
            # Fall back to the first sheet so downstream code still raises a
            # helpful "no header found" error instead of "empty file".
            first_name = next(iter(sheets))
            return first_name, sheets[first_name]

        if len(sheets) > 1:
            other_sheets = ", ".join(n for n in sheets if n != best_name)
            warnings.append(
                f"XLSX hat {len(sheets)} Sheets — verwende '{best_name}' "
                f"(beste Tabellenstruktur). Ignoriert: {other_sheets}"
            )

        return best_name, best_rows

    @staticmethod
    def _auto_detect_columns(
        headers: Sequence[str], warnings: list[str]
    ) -> dict[str, str]:
        """Pick the best header for each canonical field.

        Collisions (two headers competing for the same field) are resolved
        silently — higher-scoring wins. The chosen mapping is fully visible
        in ``detected_columns`` so the user can verify without us spamming
        the warnings panel with "X collided with Y, Y won" messages.
        """
        best_for_field: dict[str, tuple[str, float]] = {}
        for header in headers:
            if not header:
                continue
            canonical, score = fuzzy_match_canonical(header)
            if canonical is None or score < _AUTO_MAPPING_MIN_SCORE:
                continue
            current = best_for_field.get(canonical)
            if current is None or score > current[1]:
                best_for_field[canonical] = (header, score)
        return {
            canonical: header
            for canonical, (header, _score) in best_for_field.items()
            if canonical in KNOWN_CIRCUIT_FIELDS
        }

    @staticmethod
    def _validate_numeric_mappings(
        detected: dict[str, str],
        headers: Sequence[str],
        data_rows: Sequence[Sequence[Any]],
        warnings: list[str],
    ) -> dict[str, str]:
        """Drop numeric mappings whose data column is mostly non-numeric.

        Catches false-positive header matches where a string-typed column
        (e.g. "Einheit" with values "Std"/"Stk") is wrongly mapped to a
        numeric field. We sample up to 20 non-empty rows; if less than 50%
        of the values parse as numbers, the mapping is dropped.
        """
        kept: dict[str, str] = dict(detected)
        for canonical, source_col in list(detected.items()):
            if canonical not in _NUMERIC_FIELDS:
                continue
            try:
                col_idx = list(headers).index(source_col)
            except ValueError:
                continue
            numeric_hits = 0
            sampled = 0
            for row in data_rows[:50]:
                if sampled >= 20:
                    break
                if col_idx >= len(row):
                    continue
                cell = row[col_idx]
                if cell is None or str(cell).strip() == "":
                    continue
                sampled += 1
                num, _warn = _convert_value(canonical, cell, None)
                if isinstance(num, (int, float)):
                    numeric_hits += 1
            if sampled >= 3 and numeric_hits / sampled < 0.5:
                warnings.append(
                    f"Spalte '{source_col}' wurde als '{canonical}' erkannt, "
                    f"enthaelt aber nur {numeric_hits}/{sampled} numerische "
                    "Werte — Mapping verworfen."
                )
                kept.pop(canonical, None)
        return kept
