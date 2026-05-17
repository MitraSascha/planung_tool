import json
import shutil
from pathlib import Path
from typing import Any

from sqlalchemy.orm import Session

from app.core.settings import settings
from app.services.pii_tokenizer import pii_tokenizer

TEXT_SUFFIXES = {".csv", ".json", ".md", ".txt", ".html"}
EXTRACTABLE_SUFFIXES = {".pdf", ".xlsx", ".xls"}


def prepare_sanitized_generator_workspace(db: Session, slug: str) -> Path:
    workspace_path = settings.workspaces_path / slug
    generator_path = workspace_path / "generator_input"
    sanitized_docs_path = generator_path / "docs"
    source_docs_path = workspace_path / "docs"
    source_input_path = workspace_path / "input.json"

    if generator_path.exists():
        shutil.rmtree(generator_path)

    sanitized_docs_path.mkdir(parents=True, exist_ok=True)
    (generator_path / "output").mkdir(parents=True, exist_ok=True)

    manifest: dict[str, list[dict[str, str]]] = {
        "tokenized_files": [],
        "excluded_files": [],
    }

    if source_input_path.exists():
        _tokenize_file(db, source_input_path, generator_path / "input.json", slug, manifest)

    source_heating_path = workspace_path / "heating_design.json"
    if source_heating_path.exists():
        _tokenize_file(db, source_heating_path, generator_path / "heating_design.json", slug, manifest)

    # Angebots-JSON (heute eingebaut): Lieferantenname, Anschriften und
    # Item-Beschreibungen können PII enthalten — selbe Tokenisierung wie
    # heating_design.json.
    source_offers_path = workspace_path / "offers.json"
    if source_offers_path.exists():
        _tokenize_file(db, source_offers_path, generator_path / "offers.json", slug, manifest)

    # Sprachnotizen (Phase 13.3): transkribierte Voice-Notes koennen PII enthalten
    # (Kundennamen, Adressen, Telefonnummern), darum identische Tokenisierung
    # wie input.json/heating_design.json.
    source_voice_notes_path = workspace_path / "voice_notes.json"
    if source_voice_notes_path.exists():
        _tokenize_file(
            db,
            source_voice_notes_path,
            generator_path / "voice_notes.json",
            slug,
            manifest,
        )

    # Foto-Doku (Phase 12.4): ``photos.json`` enthaelt Captions/GPS, die PII
    # sein koennen — also durch den Tokenizer schicken. Die Bild-Files
    # selbst werden NICHT tokenisiert (Pixel haben kein Token-Konzept), aber
    # 1:1 in den Generator-Workspace gespiegelt, damit die KI relative Pfade
    # auf ``photos/<filename>`` setzen kann.
    source_photos_manifest = workspace_path / "photos.json"
    if source_photos_manifest.exists():
        _tokenize_file(
            db,
            source_photos_manifest,
            generator_path / "photos.json",
            slug,
            manifest,
        )

    source_photos_dir = workspace_path / "photos"
    if source_photos_dir.exists() and source_photos_dir.is_dir():
        target_photos_dir = generator_path / "photos"
        target_photos_dir.mkdir(parents=True, exist_ok=True)
        for photo_file in sorted(
            item for item in source_photos_dir.iterdir() if item.is_file()
        ):
            shutil.copy2(photo_file, target_photos_dir / photo_file.name)

    if source_docs_path.exists():
        for source_file in sorted(item for item in source_docs_path.rglob("*") if item.is_file()):
            relative_path = source_file.relative_to(source_docs_path)
            if source_file.suffix.lower() in TEXT_SUFFIXES:
                target_file = sanitized_docs_path / relative_path
                _tokenize_file(db, source_file, target_file, slug, manifest)
            elif source_file.suffix.lower() in EXTRACTABLE_SUFFIXES:
                target_file = sanitized_docs_path / f"{relative_path.as_posix()}.txt"
                try:
                    extracted_text = _extract_text(source_file)
                except Exception as exc:
                    manifest["excluded_files"].append(
                        {
                            "path": relative_path.as_posix(),
                            "reason": f"Text-Extraktion fehlgeschlagen: {exc}",
                        }
                    )
                    continue

                _tokenize_text(
                    db=db,
                    text=extracted_text,
                    target_file=target_file,
                    slug=slug,
                    scope_name=relative_path.as_posix(),
                    manifest=manifest,
                    source_path=relative_path.as_posix(),
                    extraction="extracted",
                )
            else:
                manifest["excluded_files"].append(
                    {
                        "path": relative_path.as_posix(),
                        "reason": "Dateityp wird noch nicht textuell extrahiert und deshalb nicht an Codex uebergeben.",
                    }
                )

    (generator_path / "privacy_manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return generator_path


def sync_generator_output(slug: str) -> None:
    workspace_path = settings.workspaces_path / slug
    generator_output = workspace_path / "generator_input" / "output"
    workspace_output = workspace_path / "output"

    if not generator_output.exists():
        return

    if workspace_output.exists():
        shutil.rmtree(workspace_output)

    shutil.copytree(generator_output, workspace_output)


def _tokenize_file(
    db: Session,
    source_file: Path,
    target_file: Path,
    slug: str,
    manifest: dict[str, list[dict[str, str]]],
) -> None:
    try:
        text = source_file.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        text = source_file.read_text(encoding="latin-1")

    _tokenize_text(
        db=db,
        text=text,
        target_file=target_file,
        slug=slug,
        scope_name=source_file.name,
        manifest=manifest,
        source_path=source_file.name,
        extraction="direct_text",
    )


def _tokenize_text(
    db: Session,
    text: str,
    target_file: Path,
    slug: str,
    scope_name: str,
    manifest: dict[str, list[dict[str, str]]],
    source_path: str,
    extraction: str,
) -> None:
    target_file.parent.mkdir(parents=True, exist_ok=True)

    run, sanitized_text = pii_tokenizer.tokenize(
        db=db,
        text=text,
        scope=f"project:{slug}:{scope_name}",
        mode="internal",
    )
    target_file.write_text(sanitized_text, encoding="utf-8")
    manifest["tokenized_files"].append(
        {
            "path": source_path,
            "sanitized_path": target_file.name,
            "extraction": extraction,
            "run_id": run.run_id,
        }
    )


def _extract_text(source_file: Path) -> str:
    suffix = source_file.suffix.lower()
    if suffix == ".pdf":
        return _extract_pdf_text(source_file)
    if suffix == ".xlsx":
        return _extract_xlsx_text(source_file)
    if suffix == ".xls":
        return _extract_xls_text(source_file)
    raise ValueError(f"Unsupported extractable file type: {suffix}")


def _extract_pdf_text(source_file: Path) -> str:
    from pypdf import PdfReader

    reader = PdfReader(str(source_file))
    pages: list[str] = []
    for index, page in enumerate(reader.pages, start=1):
        page_text = page.extract_text() or ""
        pages.append(f"--- Seite {index} ---\n{page_text.strip()}")
    return "\n\n".join(pages).strip()


def _extract_xlsx_text(source_file: Path) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(source_file, data_only=True, read_only=True)
    try:
        sheets: list[str] = []
        for sheet in workbook.worksheets:
            rows = [_format_row(row) for row in sheet.iter_rows(values_only=True)]
            rows = [row for row in rows if row]
            sheets.append(f"--- Tabelle: {sheet.title} ---\n" + "\n".join(rows))
        return "\n\n".join(sheets).strip()
    finally:
        workbook.close()


def _extract_xls_text(source_file: Path) -> str:
    import xlrd

    workbook = xlrd.open_workbook(str(source_file))
    sheets: list[str] = []
    for sheet in workbook.sheets():
        rows = []
        for row_index in range(sheet.nrows):
            row = [sheet.cell_value(row_index, column_index) for column_index in range(sheet.ncols)]
            formatted = _format_row(row)
            if formatted:
                rows.append(formatted)
        sheets.append(f"--- Tabelle: {sheet.name} ---\n" + "\n".join(rows))
    return "\n\n".join(sheets).strip()


def _format_row(row: tuple[Any, ...] | list[Any]) -> str:
    values = [_format_cell(value) for value in row]
    while values and values[-1] == "":
        values.pop()
    return "\t".join(values).strip()


def _format_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()
