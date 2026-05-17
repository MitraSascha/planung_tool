"""Generic xlsx/csv importer for supplier offers (Angebote).

Mirrors the structure of the heating ``generic_table`` importer but uses a
different synonym table and CORE gate (at least one of unit_price_net_eur
or total_net_eur must map, otherwise the file is not an offer).
"""
from __future__ import annotations

import csv
import difflib
import io
import re
import unicodedata
from typing import Any, Sequence

from app.models.offers import (
    OfferBase,
    OfferImportPreview,
    OfferItemBase,
)
from app.services.offer_importers.base import (
    CORE_REQUIRED_FIELDS,
    KNOWN_ITEM_FIELDS,
    OfferColumnMapping,
    OfferImporter,
    OfferImporterError,
)


_FIELD_SYNONYMS: dict[str, tuple[str, ...]] = {
    "position_label": ("position", "positionsnr", "positionsnummer", "lfdnr", "ordnung"),
    "article_no": (
        "artikelnummer",
        "artikelnr",
        "artno",
        "matnr",
        "materialnummer",
        "artikelcode",
    ),
    "name": (
        "artikelbezeichnung",
        "artikelname",
        "produktname",
        "produktbezeichnung",
        "kurzbezeichnung",
        "name",
        "artikel",
    ),
    "description": (
        "beschreibung",
        "langtext",
        "produktbeschreibung",
        "details",
        "kurztext",
        "artikeltext",
    ),
    "qty": ("menge", "anzahl", "stueckzahl", "quantity"),
    "unit": ("einheit", "mengeneinheit", "einh"),
    "unit_price_net_eur": (
        "einheitspreis",
        "stueckpreis",
        "verkaufspreis",
        "preisproeinheit",
        "epnetto",
        "einzelpreis",
    ),
    "total_net_eur": (
        "gesamtpreis",
        "positionspreis",
        "betragnetto",
        "positionsbetrag",
        "summepreis",
        "gesamtnetto",
        "nettogesamt",
        "positionspreis",
    ),
    "vat_rate": (
        "mwst",
        "mehrwertsteuer",
        "ust",
        "umsatzsteuer",
        "steuersatz",
    ),
    "notes": ("bemerkung", "anmerkung", "kommentar", "notiz"),
}

_NUMERIC_FIELDS: set[str] = {
    "qty",
    "unit_price_net_eur",
    "total_net_eur",
    "vat_rate",
}

_AUTO_MAPPING_MIN_SCORE: float = 0.78


# ----------------------------------------------------------------------
# Header normalisation + fuzzy match (kept local to this module to avoid
# coupling with the heating importer; same idea, different synonym table).
# ----------------------------------------------------------------------


def normalize_header(header: str | None) -> str:
    if header is None:
        return ""
    text = str(header).strip().lower()
    text = re.sub(r"[\[\(\{][^\]\)\}]*[\]\)\}]", " ", text)
    text = "".join(ch for ch in text if ch.isprintable())
    for src, dst in {"ä": "a", "ö": "o", "ü": "u", "ß": "ss"}.items():
        text = text.replace(src, dst)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", "", text).strip()


def fuzzy_match_canonical(header: str | None) -> tuple[str | None, float]:
    normalized = normalize_header(header)
    if not normalized:
        return None, 0.0
    best_field: str | None = None
    best_score = 0.0
    for canonical, synonyms in _FIELD_SYNONYMS.items():
        for synonym in synonyms:
            if not synonym:
                continue
            if synonym == normalized:
                score = 1.0
            elif len(synonym) >= 4 and synonym in normalized:
                score = max(0.85, len(synonym) / max(len(normalized), 1))
            else:
                score = difflib.SequenceMatcher(None, normalized, synonym).ratio()
            if score > best_score:
                best_score = score
                best_field = canonical
    if best_score < 0.5:
        return None, best_score
    return best_field, best_score


# ----------------------------------------------------------------------
# Value parsing.
# ----------------------------------------------------------------------


def _to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).strip().replace(" ", "").replace("\xa0", "")
    cleaned = cleaned.replace("€", "").replace("EUR", "").replace("%", "")
    if not cleaned:
        return None
    if "," in cleaned and "." in cleaned:
        if cleaned.rfind(",") > cleaned.rfind("."):
            cleaned = cleaned.replace(".", "").replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
    else:
        cleaned = cleaned.replace(",", ".")
    try:
        return float(cleaned)
    except (TypeError, ValueError):
        return None


def _coerce_value(field: str, value: Any) -> tuple[Any, str | None]:
    if value is None or value == "":
        return None, None
    if field in _NUMERIC_FIELDS:
        num = _to_float(value)
        if num is None:
            return None, f"Wert nicht numerisch fuer Feld '{field}': {value!r}"
        return num, None
    return str(value).strip(), None


# ----------------------------------------------------------------------
# File-format handling.
# ----------------------------------------------------------------------


def _decode_csv_bytes(content: bytes) -> str:
    if content[:3] == b"\xef\xbb\xbf":
        return content[3:].decode("utf-8", errors="replace")
    for encoding in ("utf-8", "cp1252", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def _sniff_csv(text: str) -> csv.Dialect:
    sample = text[:4096]
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        first = sample.splitlines()[:5]
        semi = sum(line.count(";") for line in first)
        comma = sum(line.count(",") for line in first)
        tab = sum(line.count("\t") for line in first)
        if tab >= semi and tab >= comma and tab > 0:
            return csv.excel_tab  # type: ignore[return-value]
        if semi > comma:
            class _Semi(csv.excel):
                delimiter = ";"

            return _Semi()  # type: ignore[return-value]
        return csv.excel()  # type: ignore[return-value]


def _read_csv_rows(content: bytes) -> list[list[str]]:
    text = _decode_csv_bytes(content)
    if not text.strip():
        return []
    dialect = _sniff_csv(text)
    reader = csv.reader(io.StringIO(text), dialect=dialect)
    rows: list[list[str]] = []
    for raw in reader:
        row = [cell.strip() if isinstance(cell, str) else cell for cell in raw]
        while row and (row[-1] is None or row[-1] == ""):
            row.pop()
        rows.append(row)
    return rows


def _read_xlsx_rows(content: bytes) -> list[list[Any]]:
    try:
        import openpyxl  # type: ignore
    except ImportError as exc:  # pragma: no cover
        raise OfferImporterError(
            "openpyxl ist nicht installiert; .xlsx kann nicht gelesen werden"
        ) from exc
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        raise OfferImporterError(f"Excel-Datei konnte nicht geoeffnet werden: {exc}") from exc
    sheet = workbook[workbook.sheetnames[0]]
    rows: list[list[Any]] = []
    for raw in sheet.iter_rows(values_only=True):
        row = list(raw)
        while row and (row[-1] is None or row[-1] == ""):
            row.pop()
        rows.append(row)
    workbook.close()
    return rows


# ----------------------------------------------------------------------
# Header detection.
# ----------------------------------------------------------------------


def _row_best_per_field(row: Sequence[Any]) -> dict[str, float]:
    best: dict[str, float] = {}
    for cell in row:
        if cell is None:
            continue
        text = str(cell).strip()
        if not text:
            continue
        canonical, score = fuzzy_match_canonical(text)
        if canonical is None or score < 0.5:
            continue
        if score > best.get(canonical, 0.0):
            best[canonical] = score
    return best


def _find_header_row(rows: Sequence[Sequence[Any]]) -> int | None:
    best_idx: int | None = None
    best_strong = 0
    best_weak = 0
    # Scan first 30 rows — offer headers can sit below an Anschreiben block.
    for idx, row in enumerate(rows[:30]):
        per = _row_best_per_field(row)
        if not per:
            continue
        strong = sum(1 for s in per.values() if s >= 0.85)
        weak = len(per)
        if strong < 2:
            continue
        if (strong, weak) > (best_strong, best_weak):
            best_strong = strong
            best_weak = weak
            best_idx = idx
    return best_idx


# ----------------------------------------------------------------------
# Importer.
# ----------------------------------------------------------------------


class GenericOfferTableImporter(OfferImporter):
    source_name = "generic_table"
    display_name = "Excel / CSV (Angebote)"
    accepts_extensions = (".xlsx", ".xls", ".csv")

    def can_handle(self, filename: str, content_head: bytes) -> bool:
        lower = filename.lower()
        return any(lower.endswith(ext) for ext in self.accepts_extensions)

    def parse(
        self,
        filename: str,
        content: bytes,
        mapping: OfferColumnMapping | None = None,
    ) -> OfferImportPreview:
        if not content:
            raise OfferImporterError("Datei ist leer.")
        lower = filename.lower()
        if lower.endswith(".xlsx"):
            rows = _read_xlsx_rows(content)
        elif lower.endswith(".csv"):
            rows = _read_csv_rows(content)
        else:
            # .xls — try xlrd via the heating importer's helper if needed.
            try:
                import xlrd  # type: ignore
            except ImportError as exc:
                raise OfferImporterError(
                    ".xls wird (noch) nicht unterstuetzt — bitte als .xlsx oder .csv exportieren."
                ) from exc
            try:
                book = xlrd.open_workbook(file_contents=content)
            except Exception as exc:
                raise OfferImporterError(f".xls konnte nicht gelesen werden: {exc}") from exc
            sheet = book.sheet_by_index(0)
            rows = []
            for i in range(sheet.nrows):
                row = list(sheet.row_values(i))
                while row and (row[-1] is None or row[-1] == ""):
                    row.pop()
                rows.append(row)

        rows = [r for r in rows if any(c is not None and str(c).strip() != "" for c in r)]
        if not rows:
            raise OfferImporterError("Datei enthaelt keine lesbaren Zeilen.")

        header_idx = _find_header_row(rows)
        if header_idx is None:
            raise OfferImporterError(
                "Keine plausible Header-Zeile fuer ein Angebot gefunden. "
                "Benoetigt werden mindestens zwei stark gematchte Spalten "
                "(z.B. 'Artikelnummer', 'Menge', 'Einheitspreis', 'Gesamtpreis')."
            )

        meta_rows = rows[:header_idx]
        header_row = rows[header_idx]
        data_rows = rows[header_idx + 1:]

        headers: list[str] = [str(c).strip() if c is not None else "" for c in header_row]
        warnings: list[str] = []

        detected = self._auto_detect_columns(headers, warnings)
        manual_override = mapping is not None and bool(mapping.item_columns)
        if manual_override:
            valid: dict[str, str] = {}
            for canonical, source_col in mapping.item_columns.items():
                if source_col in headers:
                    valid[canonical] = source_col
                else:
                    warnings.append(
                        f"Manuelles Mapping: Spalte '{source_col}' nicht in Datei gefunden"
                    )
            detected = valid

        if not manual_override:
            detected = self._validate_numeric_mappings(detected, headers, data_rows, warnings)

        # CORE-Gate: ohne mind. einen Preis ist es kein Angebot.
        if not manual_override and not any(f in detected for f in CORE_REQUIRED_FIELDS):
            raise OfferImporterError(
                "Diese Datei enthaelt kein erkennbares Angebot: Es konnte "
                "weder eine Spalte fuer 'Einheitspreis' noch fuer "
                "'Gesamtpreis' gefunden werden. Bitte pruefen Sie, ob Sie die "
                "richtige Datei (Angebot / Stueckliste) hochgeladen haben. "
                "Erkannte Spalten: " + (", ".join(h for h in headers if h) or "(keine)")
            )

        items, warns_added = self._build_items(headers, data_rows, detected)
        warnings.extend(warns_added)
        if not items:
            raise OfferImporterError(
                "Keine Positionen erkannt — eventuell ist die Datei leer "
                "oder die Spalten passen nicht."
            )

        offer_meta = self._extract_offer_meta(filename, meta_rows, items)

        # Surface unmapped headers as a single summary line instead of one
        # warning per column — "Lieferanten-Nr / EAN / MwSt." is just noise
        # individually but the aggregate is useful info for the user.
        mapped = set(detected.values())
        unmapped_headers = [h for h in headers if h and h not in mapped]
        if unmapped_headers:
            warnings.append(
                "Nicht ins Angebots-Modell übernommen: "
                + ", ".join(unmapped_headers)
            )

        source_type: str = "xlsx" if lower.endswith((".xlsx", ".xls")) else "csv"
        return OfferImportPreview(
            source_type=source_type,  # type: ignore[arg-type]
            source_file=filename,
            offer=offer_meta,
            items=items,
            warnings=warnings,
            detected_columns=detected,
        )

    # ------------------------------------------------------------------

    @staticmethod
    def _auto_detect_columns(
        headers: Sequence[str], warnings: list[str]
    ) -> dict[str, str]:
        """Pick the best header for each canonical field.

        Collisions (two headers competing for the same canonical field) are
        resolved silently — the higher-scoring header wins. We don't log
        these as warnings because they're noise: the user doesn't need to
        know that 'Einkaufspreis' also matched 'unit_price_net_eur' if
        'Einheitspreis' won. The chosen mapping is visible in
        ``detected_columns`` for full transparency.
        """
        best_for_field: dict[str, tuple[str, float]] = {}
        for header in headers:
            if not header:
                continue
            canonical, score = fuzzy_match_canonical(header)
            if canonical is None or score < _AUTO_MAPPING_MIN_SCORE:
                continue
            current = best_for_field.get(canonical)
            if current is None or score > current[1]:
                best_for_field[canonical] = (header, score)
        return {
            canonical: header
            for canonical, (header, _s) in best_for_field.items()
            if canonical in KNOWN_ITEM_FIELDS
        }

    @staticmethod
    def _validate_numeric_mappings(
        detected: dict[str, str],
        headers: Sequence[str],
        data_rows: Sequence[Sequence[Any]],
        warnings: list[str],
    ) -> dict[str, str]:
        kept = dict(detected)
        for canonical, source_col in list(detected.items()):
            if canonical not in _NUMERIC_FIELDS:
                continue
            try:
                col_idx = list(headers).index(source_col)
            except ValueError:
                continue
            numeric_hits = 0
            sampled = 0
            for row in data_rows[:50]:
                if sampled >= 20:
                    break
                if col_idx >= len(row):
                    continue
                cell = row[col_idx]
                if cell is None or str(cell).strip() == "":
                    continue
                sampled += 1
                if _to_float(cell) is not None:
                    numeric_hits += 1
            if sampled >= 3 and numeric_hits / sampled < 0.5:
                warnings.append(
                    f"Spalte '{source_col}' wurde als '{canonical}' erkannt, "
                    f"enthaelt aber nur {numeric_hits}/{sampled} numerische "
                    "Werte — Mapping verworfen."
                )
                kept.pop(canonical, None)
        return kept

    @staticmethod
    def _build_items(
        headers: Sequence[str],
        data_rows: Sequence[Sequence[Any]],
        detected: dict[str, str],
    ) -> tuple[list[OfferItemBase], list[str]]:
        canonical_idx: list[tuple[str, int]] = []
        for canonical, col in detected.items():
            try:
                canonical_idx.append((canonical, list(headers).index(col)))
            except ValueError:
                continue
        warnings: list[str] = []
        seen_warns: set[str] = set()
        items: list[OfferItemBase] = []
        for position, row in enumerate(data_rows):
            if not any(c is not None and str(c).strip() != "" for c in row):
                continue
            values: dict[str, Any] = {"position_index": position}
            for canonical, idx in canonical_idx:
                raw = row[idx] if idx < len(row) else None
                coerced, warn = _coerce_value(canonical, raw)
                if warn and warn not in seen_warns:
                    warnings.append(warn)
                    seen_warns.add(warn)
                if coerced is not None:
                    values[canonical] = coerced
            non_pos = {k: v for k, v in values.items() if k != "position_index"}
            if not non_pos:
                continue
            try:
                items.append(OfferItemBase(**values))
            except Exception as exc:  # pragma: no cover - pydantic is permissive
                warnings.append(f"Zeile {position + 1} uebersprungen: {exc}")
        return items, warnings

    @staticmethod
    def _extract_offer_meta(
        filename: str,
        meta_rows: Sequence[Sequence[Any]],
        items: Sequence[OfferItemBase],
    ) -> OfferBase:
        """Pull supplier / offer-no / date out of the rows above the header.

        Falls back to filename parsing (e.g. ``Angebot-ANG-614-IMMOMAKS-...``)
        when meta-rows are empty.
        """
        supplier_name: str | None = None
        offer_no: str | None = None
        for row in meta_rows:
            for cell in row:
                if cell is None:
                    continue
                text = str(cell).strip()
                if not text:
                    continue
                norm = normalize_header(text)
                if "lieferant" in norm or "anbieter" in norm or "firma" in norm:
                    parts = re.split(r"[:=]", text, maxsplit=1)
                    if len(parts) == 2 and parts[1].strip():
                        supplier_name = parts[1].strip()
                if "angebotsnr" in norm or "angebotnr" in norm:
                    parts = re.split(r"[:=]", text, maxsplit=1)
                    if len(parts) == 2 and parts[1].strip():
                        offer_no = parts[1].strip()

        if supplier_name is None or offer_no is None:
            base = filename.rsplit("/", 1)[-1]
            base_no_ext = base.rsplit(".", 1)[0]
            # Pattern e.g. "Angebot-ANG-614-IMMOMAKS-..." → offer_no=ANG-614, supplier=IMMOMAKS
            match = re.match(
                r"(?i)Angebot-(?P<no>[A-Z]+-\d+)[ _-](?P<supplier>[A-Za-z0-9_]+)",
                base_no_ext,
            )
            if match:
                if offer_no is None:
                    offer_no = match.group("no")
                if supplier_name is None:
                    supplier_name = match.group("supplier")

        total_net = sum(
            (it.total_net_eur for it in items if it.total_net_eur is not None), 0.0
        ) or None

        return OfferBase(
            supplier_name=supplier_name or "Unbekannter Lieferant",
            offer_no=offer_no,
            offer_date=None,
            currency="EUR",
            total_net_eur=total_net,
            total_gross_eur=None,
            vat_rate=None,
            notes=None,
        )
