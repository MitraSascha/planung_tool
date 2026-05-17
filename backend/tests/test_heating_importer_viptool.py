"""Tests for ``app.services.heating_importers.viptool_master.ViptoolMasterImporter``.

Phase 11.2: bis eine echte Viega-Viptool-Beispieldatei vorliegt, deckt dieser
Test programmatisch erzeugte Mini-XLSX-Dateien ab, die das oeffentlich
dokumentierte Spaltenlayout abbilden. Sobald ein realer Export verfuegbar
ist, werden hier zusaetzliche Fixture-Tests ergaenzt.
"""
from __future__ import annotations

import io

import pytest
from openpyxl import Workbook

from app.services.heating_importers import detect_importer
from app.services.heating_importers.base import HeatingImporterError
from app.services.heating_importers.viptool_master import ViptoolMasterImporter


# ---------------------------------------------------------------------------
# Helpers — synthetic Viptool-shaped workbooks
# ---------------------------------------------------------------------------


def _viptool_workbook_bytes(
    *,
    sheet_title: str = "Strangberechnung",
    heat_load_unit: str = "W",
    pressure_unit: str = "Pa",
    flow_unit: str = "l/h",
    rows: list[tuple] | None = None,
    brand_marker_in_a1: bool = False,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    if brand_marker_in_a1:
        ws["A1"] = "Viega Viptool Master — Strangberechnung Export"

    header_row_idx = 3 if brand_marker_in_a1 else 1

    headers = [
        "Strang",
        "Raum",
        "Geschoss",
        "Heizflaeche",
        f"Norm-Heizlast Q [{heat_load_unit}]",
        f"Volumenstrom [{flow_unit}]",
        f"Druckverlust [{pressure_unit}]",
        "Rohrlaenge gesamt [m]",
        "Ventil",
        "Voreinstellung",
        "kv-Wert",
    ]
    for col, value in enumerate(headers, start=1):
        ws.cell(row=header_row_idx, column=col, value=value)

    default_rows = rows if rows is not None else [
        ("S1", "Wohnzimmer", "EG", "HK-22-1000", 1450, 62, 85, 8.5, "Heimeier", "4", 0.42),
        ("S1", "Kueche",     "EG", "HK-22-600",  900,  38, 60, 5.0, "Heimeier", "3", 0.30),
        ("S2", "Bad",        "OG", "HK-33-800",  1100, 47, 70, 6.2, "Heimeier", "5", 0.55),
    ]
    for offset, row in enumerate(default_rows, start=1):
        for col, value in enumerate(row, start=1):
            ws.cell(row=header_row_idx + offset, column=col, value=value)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# can_handle
# ---------------------------------------------------------------------------


def test_can_handle_filename_with_viptool() -> None:
    importer = ViptoolMasterImporter()
    content = _viptool_workbook_bytes()
    assert importer.can_handle("viptool_export.xlsx", content) is True


def test_can_handle_rejects_unrelated_xlsx() -> None:
    """A generic-looking xlsx without Viptool brand markers must NOT be
    claimed — that's the GenericTableImporter's job."""
    importer = ViptoolMasterImporter()
    # Synthesise a workbook that has neither "viptool" in the filename nor in A1.
    wb = Workbook()
    ws = wb.active
    ws.title = "Tabelle1"
    ws["A1"] = "Strang"
    ws["B1"] = "Raum"
    buf = io.BytesIO()
    wb.save(buf)
    assert importer.can_handle("strangliste.xlsx", buf.getvalue()) is False


def test_can_handle_detects_brand_in_first_row() -> None:
    importer = ViptoolMasterImporter()
    content = _viptool_workbook_bytes(brand_marker_in_a1=True)
    # Even without "viptool" in the filename the brand marker in A1 triggers.
    assert importer.can_handle("strangberechnung_2026.xlsx", content) is True


def test_can_handle_rejects_non_xlsx_extensions() -> None:
    importer = ViptoolMasterImporter()
    assert importer.can_handle("viptool.csv", b"anything") is False
    assert importer.can_handle("viptool.pdf", b"anything") is False


# ---------------------------------------------------------------------------
# Registry routing — detect_importer must pick Viptool first when applicable
# ---------------------------------------------------------------------------


def test_registry_selects_viptool_for_branded_filename() -> None:
    content = _viptool_workbook_bytes()
    importer = detect_importer("viptool_export.xlsx", content)
    assert importer is not None
    assert isinstance(importer, ViptoolMasterImporter)


# ---------------------------------------------------------------------------
# parse — happy path + unit conversion
# ---------------------------------------------------------------------------


def test_parse_extracts_circuits_with_default_mapping() -> None:
    importer = ViptoolMasterImporter()
    content = _viptool_workbook_bytes()
    preview = importer.parse("viptool_export.xlsx", content)

    assert preview.source == "viptool_xlsx"
    assert preview.source_file == "viptool_export.xlsx"
    assert len(preview.circuits) == 3

    first = preview.circuits[0]
    assert first.strand == "S1"
    assert first.room == "Wohnzimmer"
    assert first.floor == "EG"
    assert first.heat_load_w == pytest.approx(1450.0)
    assert first.volume_flow_lph == pytest.approx(62.0)
    assert first.pressure_drop_pa == pytest.approx(85.0)
    assert first.pipe_length_m == pytest.approx(8.5)
    assert first.valve_preset == "4"
    assert first.kv_value == pytest.approx(0.42)


def test_parse_converts_kw_to_watts() -> None:
    importer = ViptoolMasterImporter()
    # Heizlast in kW statt W — Adapter muss die Einheit aus dem Header lesen
    # und korrekt nach Watt konvertieren.
    content = _viptool_workbook_bytes(
        heat_load_unit="kW",
        rows=[
            ("S1", "Wohnzimmer", "EG", "HK-22", 1.45, 62, 85, 8.5, "Heimeier", "4", 0.42),
            ("S1", "Kueche",     "EG", "HK-22", 0.9,  38, 60, 5.0, "Heimeier", "3", 0.30),
        ],
    )
    preview = importer.parse("viptool_export.xlsx", content)
    assert preview.circuits[0].heat_load_w == pytest.approx(1450.0)
    assert preview.circuits[1].heat_load_w == pytest.approx(900.0)


def test_parse_converts_mbar_to_pascal() -> None:
    importer = ViptoolMasterImporter()
    content = _viptool_workbook_bytes(
        pressure_unit="mbar",
        rows=[
            ("S1", "Wohnzimmer", "EG", "HK-22", 1450, 62, 0.85, 8.5, "Heimeier", "4", 0.42),
        ],
    )
    preview = importer.parse("viptool_export.xlsx", content)
    # 0.85 mbar -> 85 Pa
    assert preview.circuits[0].pressure_drop_pa == pytest.approx(85.0)


def test_parse_emits_stub_warning() -> None:
    importer = ViptoolMasterImporter()
    content = _viptool_workbook_bytes()
    preview = importer.parse("viptool_export.xlsx", content)
    assert any(
        "Default-Stub" in w or "Beispieldatei" in w for w in preview.warnings
    ), preview.warnings


def test_parse_design_stays_none_when_no_overrides() -> None:
    """Viptool-Exports liefern keine Anlagenkenndaten in der Kreis-Tabelle —
    der Adapter MUSS sie als None lassen (keine Warnung); der Generator
    weist sie spaeter als 'Offene Punkte' aus."""
    importer = ViptoolMasterImporter()
    content = _viptool_workbook_bytes()
    preview = importer.parse("viptool_export.xlsx", content)
    assert preview.design.supply_temp_c is None
    assert preview.design.return_temp_c is None
    assert preview.design.total_volume_flow_lph is None


# ---------------------------------------------------------------------------
# parse — error paths
# ---------------------------------------------------------------------------


def test_parse_raises_on_corrupt_file() -> None:
    importer = ViptoolMasterImporter()
    with pytest.raises(HeatingImporterError):
        importer.parse("viptool_export.xlsx", b"\x00 not a real xlsx \x00")


def test_parse_raises_when_no_header_found() -> None:
    importer = ViptoolMasterImporter()
    wb = Workbook()
    ws = wb.active
    ws.title = "Strangberechnung"
    ws["A1"] = "irgendwas"
    ws["B1"] = "anderes"
    buf = io.BytesIO()
    wb.save(buf)
    with pytest.raises(HeatingImporterError):
        importer.parse("viptool_export.xlsx", buf.getvalue())


def test_parse_raises_when_only_header_row() -> None:
    importer = ViptoolMasterImporter()
    content = _viptool_workbook_bytes(rows=[])
    with pytest.raises(HeatingImporterError):
        importer.parse("viptool_export.xlsx", content)
